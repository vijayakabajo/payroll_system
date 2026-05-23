"""
Bulk-upload views: upload Excel → parse → process each row via
PayrollEngine → show progress → history list.
"""

import csv
import io
import logging
import os

from django.contrib import messages
from django.core.files.base import ContentFile
from django.core.paginator import Paginator
from django.http import HttpResponse, Http404
from django.shortcuts import get_object_or_404, redirect, render

from openpyxl import Workbook

from apps.employees.models import Employee

from .forms import BulkUploadForm
from .models import BulkUpload, BulkUploadRow
from .parser import (
    ExcelParser,
    TEMPLATE_COLUMN_ORDER,
    EARNING_COLUMNS,
    DEDUCTION_COLUMNS,
    LEAVE_COLUMNS,
)

logger = logging.getLogger(__name__)

# Try importing PayrollEngine — it may not exist yet.
try:
    from apps.payroll.engine import PayrollEngine
except ImportError:
    PayrollEngine = None


# ─────────────────────────────────────────────────────────────
#  PROCESSING LOGIC
# ─────────────────────────────────────────────────────────────

def _build_overrides(row: dict) -> dict:
    """
    Convert a parsed row dict into the ``overrides`` dict expected by
    ``PayrollEngine.generate_payslip()``.
    """
    overrides = {}
    for key in EARNING_COLUMNS | DEDUCTION_COLUMNS:
        value = row.get(key)
        if value is not None:
            overrides[key] = value
    return overrides


def _build_leave_data(row: dict) -> dict:
    """Extract leave-related fields from the row."""
    leave_data = {}
    for key in LEAVE_COLUMNS:
        value = row.get(key)
        if value is not None:
            leave_data[key] = value
    return leave_data


def _process_upload(bulk_upload: BulkUpload):
    """
    Run the full processing pipeline for *bulk_upload*:
    1. Parse the Excel file.
    2. Create BulkUploadRow records (valid + invalid).
    3. For each valid row, call PayrollEngine.
    4. Update counters and final status.
    5. Generate an error-report file for any failures.
    """

    bulk_upload.status = 'processing'
    bulk_upload.save(update_fields=['status'])

    parser = ExcelParser()
    bulk_upload.file.seek(0)
    valid_rows, parse_errors = parser.parse(bulk_upload.file)

    # ── Create rows for parse-level errors ────────────────
    for err in parse_errors:
        BulkUploadRow.objects.create(
            bulk_upload=bulk_upload,
            row_number=err['row_number'],
            employee_code=err.get('employee_code', ''),
            raw_data=err.get('raw_data', {}),
            status='failed',
            error_message='; '.join(err['errors']),
        )

    total_rows = len(valid_rows) + len(parse_errors)
    bulk_upload.total_rows = total_rows
    bulk_upload.failed_count = len(parse_errors)
    bulk_upload.save(update_fields=['total_rows', 'failed_count'])

    success_count = 0
    fail_count = len(parse_errors)

    for row in valid_rows:
        emp_code = row['employee_code']
        row_number = row.get('_row_number', 0)
        raw_data = row.get('_raw', {})

        # Look up employee
        try:
            employee = Employee.objects.get(employee_code=emp_code)
        except Employee.DoesNotExist:
            BulkUploadRow.objects.create(
                bulk_upload=bulk_upload,
                row_number=row_number,
                employee_code=emp_code,
                raw_data=raw_data,
                status='failed',
                error_message=f'Employee with code "{emp_code}" not found.',
            )
            fail_count += 1
            bulk_upload.failed_count = fail_count
            bulk_upload.save(update_fields=['failed_count'])
            continue

        # Build overrides & leave data
        overrides = _build_overrides(row)
        leave_data = _build_leave_data(row)
        paid_days = row.get('paid_days', 30) or 30
        total_days = row.get('total_days', 30) or 30

        # Call PayrollEngine
        if PayrollEngine is None:
            BulkUploadRow.objects.create(
                bulk_upload=bulk_upload,
                row_number=row_number,
                employee_code=emp_code,
                raw_data=raw_data,
                status='failed',
                error_message='PayrollEngine is not available yet.',
            )
            fail_count += 1
            bulk_upload.failed_count = fail_count
            bulk_upload.save(update_fields=['failed_count'])
            continue

        try:
            engine = PayrollEngine()
            payslip = engine.generate_payslip(
                employee=employee,
                month=bulk_upload.month,
                year=bulk_upload.year,
                overrides=overrides or None,
                paid_days=paid_days,
                total_days=total_days,
                leave_data=leave_data or None,
            )
            BulkUploadRow.objects.create(
                bulk_upload=bulk_upload,
                row_number=row_number,
                employee_code=emp_code,
                raw_data=raw_data,
                status='success',
                payslip=payslip,
            )
            success_count += 1
            bulk_upload.success_count = success_count
            bulk_upload.save(update_fields=['success_count'])

        except Exception as exc:
            logger.exception('Payslip generation failed for %s', emp_code)
            BulkUploadRow.objects.create(
                bulk_upload=bulk_upload,
                row_number=row_number,
                employee_code=emp_code,
                raw_data=raw_data,
                status='failed',
                error_message=str(exc),
            )
            fail_count += 1
            bulk_upload.failed_count = fail_count
            bulk_upload.save(update_fields=['failed_count'])

    # ── Final status ──────────────────────────────────────
    if fail_count == 0:
        bulk_upload.status = 'completed'
    elif success_count == 0:
        bulk_upload.status = 'failed'
    else:
        bulk_upload.status = 'partial'
    bulk_upload.save(update_fields=['status'])

    # ── Generate error report ─────────────────────────────
    failed_rows = bulk_upload.rows.filter(status='failed')
    if failed_rows.exists():
        _generate_error_report(bulk_upload, failed_rows)


def _generate_error_report(bulk_upload, failed_rows):
    """Create a CSV error report and attach it to the BulkUpload."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(['Row #', 'Employee Code', 'Error Message', 'Raw Data'])

    for row in failed_rows:
        raw_str = ', '.join(f'{k}={v}' for k, v in row.raw_data.items()) if row.raw_data else ''
        writer.writerow([row.row_number, row.employee_code, row.error_message, raw_str])

    filename = f'error_report_upload_{bulk_upload.pk}.csv'
    bulk_upload.error_report.save(filename, ContentFile(buf.getvalue().encode('utf-8')), save=True)


# ─────────────────────────────────────────────────────────────
#  VIEWS
# ─────────────────────────────────────────────────────────────

def upload_view(request):
    """GET → upload form · POST → save file, process, redirect to progress."""
    if request.method == 'POST':
        form = BulkUploadForm(request.POST, request.FILES)
        if form.is_valid():
            bulk = BulkUpload.objects.create(
                file=form.cleaned_data['file'],
                month=form.cleaned_data['month'],
                year=form.cleaned_data['year'],
                uploaded_by=request.user.get_username() if request.user.is_authenticated else 'admin',
            )
            _process_upload(bulk)
            messages.success(request, f'Upload processed — {bulk.success_count} succeeded, {bulk.failed_count} failed.')
            return redirect('bulk_upload:progress', pk=bulk.pk)
    else:
        form = BulkUploadForm()

    return render(request, 'bulk_upload/upload.html', {'form': form})


def upload_progress(request, pk):
    """Show processing progress / results for a single upload."""
    bulk = get_object_or_404(BulkUpload, pk=pk)
    rows = bulk.rows.all().select_related('payslip')

    # If HTMX is requesting and the upload is still processing, return a partial
    is_htmx = request.headers.get('HX-Request') == 'true'

    return render(request, 'bulk_upload/progress.html', {
        'bulk': bulk,
        'rows': rows,
        'is_htmx': is_htmx,
    })


def upload_history(request):
    """Paginated list of all bulk uploads with optional month/year filters."""
    qs = BulkUpload.objects.all()

    month = request.GET.get('month', '')
    year = request.GET.get('year', '')

    if month:
        qs = qs.filter(month=month)
    if year:
        try:
            qs = qs.filter(year=int(year))
        except (ValueError, TypeError):
            pass

    paginator = Paginator(qs, 15)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    return render(request, 'bulk_upload/history.html', {
        'page_obj': page_obj,
        'month': month,
        'year': year,
    })


def download_template(request):
    """Generate and serve a blank .xlsx template with expected headers."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Payroll Data'

    # Write header row
    for col_idx, header in enumerate(TEMPLATE_COLUMN_ORDER, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = cell.font.copy(bold=True)

    # Adjust column widths
    for col_idx, header in enumerate(TEMPLATE_COLUMN_ORDER, start=1):
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = max(len(header) + 4, 12)

    # Write one sample row
    sample = {
        'employee_code': 'MMPL-001',
        'basic': 15000, 'hra': 7500, 'ca': 1600, 'cca': 0, 'bonus': 0,
        'medical': 1250, 'mobile': 0, 'washing': 0, 'special': 0,
        'pf': 1800, 'esi': 0, 'pt': 200, 'tds': 0, 'loan': 0, 'adv_sal': 0,
        'paid_days': 30, 'total_days': 30,
        'cl': 0, 'sl': 0, 'lwp': 0, 'ml': 0, 'pl': 0,
    }
    for col_idx, header in enumerate(TEMPLATE_COLUMN_ORDER, start=1):
        ws.cell(row=2, column=col_idx, value=sample.get(header, ''))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="payroll_upload_template.xlsx"'
    return response


def download_error_report(request, pk):
    """Serve the error report file for a specific upload."""
    bulk = get_object_or_404(BulkUpload, pk=pk)

    if not bulk.error_report:
        raise Http404('No error report available for this upload.')

    response = HttpResponse(
        bulk.error_report.read(),
        content_type='text/csv',
    )
    filename = os.path.basename(bulk.error_report.name)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
