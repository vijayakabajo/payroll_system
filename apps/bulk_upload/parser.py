"""
Excel parser for the bulk-upload payslip workflow.
Reads an .xlsx file uploaded by the user, normalises headers,
validates each row, and returns a list of dicts ready for the
payroll engine.
"""

import io
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook


# ── Column definitions ──────────────────────────────────────
# Maps normalised key → whether it is required.

REQUIRED_COLUMNS = {'employee_code'}

EARNING_COLUMNS = {
    'basic', 'hra', 'ca', 'cca', 'bonus',
    'medical', 'mobile', 'washing', 'special',
}

DEDUCTION_COLUMNS = {
    'pf', 'esi', 'pt', 'tds', 'loan', 'adv_sal',
}

ATTENDANCE_COLUMNS = {
    'paid_days', 'total_days',
}

LEAVE_COLUMNS = {
    'cl', 'sl', 'lwp', 'ml', 'pl',
}

NUMERIC_COLUMNS = EARNING_COLUMNS | DEDUCTION_COLUMNS | ATTENDANCE_COLUMNS | LEAVE_COLUMNS

ALL_EXPECTED_COLUMNS = REQUIRED_COLUMNS | NUMERIC_COLUMNS

# Pretty-print column order for template generation
TEMPLATE_COLUMN_ORDER = [
    'employee_code',
    'basic', 'hra', 'ca', 'cca', 'bonus',
    'medical', 'mobile', 'washing', 'special',
    'pf', 'esi', 'pt', 'tds', 'loan', 'adv_sal',
    'paid_days', 'total_days',
    'cl', 'sl', 'lwp', 'ml', 'pl',
]


def _normalise_header(raw: str) -> str:
    """Lower-case, strip whitespace, replace spaces/hyphens with underscores."""
    return raw.strip().lower().replace(' ', '_').replace('-', '_')


class ExcelParser:
    """
    Parse and validate an uploaded .xlsx payroll file.

    Usage::

        parser = ExcelParser()
        rows, errors = parser.parse(file_obj)
        # rows  → list[dict]  (validated row dicts)
        # errors → list[dict] (per-row validation errors)
    """

    def parse(self, file_obj):
        """
        Read the workbook, normalise headers, validate every row.

        Parameters
        ----------
        file_obj : InMemoryUploadedFile | file-like
            The uploaded .xlsx file.

        Returns
        -------
        tuple[list[dict], list[dict]]
            (valid_rows, error_rows)
            Each valid row dict has normalised keys.
            Each error dict has keys: row_number, employee_code, errors (list[str]), raw_data (dict).
        """
        try:
            wb = load_workbook(
                filename=io.BytesIO(file_obj.read()),
                read_only=True,
                data_only=True,
            )
        except Exception as exc:
            return [], [{'row_number': 0, 'employee_code': '', 'errors': [f'Cannot read Excel file: {exc}'], 'raw_data': {}}]

        ws = wb.active
        if ws is None:
            return [], [{'row_number': 0, 'employee_code': '', 'errors': ['Workbook has no active sheet.'], 'raw_data': {}}]

        rows_iter = ws.iter_rows(values_only=True)

        # ── Read header row ───────────────────────────────
        try:
            raw_headers = next(rows_iter)
        except StopIteration:
            return [], [{'row_number': 0, 'employee_code': '', 'errors': ['The uploaded file is empty.'], 'raw_data': {}}]

        headers = [_normalise_header(str(h)) if h else '' for h in raw_headers]

        # Check that employee_code column exists
        if 'employee_code' not in headers:
            return [], [{
                'row_number': 0,
                'employee_code': '',
                'errors': ['Missing required column: employee_code'],
                'raw_data': {},
            }]

        valid_rows = []
        error_rows = []

        for idx, row_values in enumerate(rows_iter, start=2):  # Excel row 2+
            # Skip entirely blank rows
            if all(v is None or str(v).strip() == '' for v in row_values):
                continue

            raw_data = {}
            for col_idx, header in enumerate(headers):
                if header and col_idx < len(row_values):
                    raw_data[header] = row_values[col_idx]

            row_errors = []
            employee_code = str(raw_data.get('employee_code', '') or '').strip()

            # Required: employee_code
            if not employee_code:
                row_errors.append('employee_code is required.')

            # Validate numeric fields
            cleaned = {'employee_code': employee_code}
            for key in NUMERIC_COLUMNS:
                val = raw_data.get(key)
                if val is None or str(val).strip() == '':
                    # Leave as None — the engine will use defaults
                    cleaned[key] = None
                    continue
                try:
                    num = Decimal(str(val).strip())
                    if num < 0:
                        row_errors.append(f'{key} must be ≥ 0 (got {val}).')
                    else:
                        cleaned[key] = float(num)
                except (InvalidOperation, ValueError):
                    row_errors.append(f'{key} must be a number (got "{val}").')

            if row_errors:
                error_rows.append({
                    'row_number': idx,
                    'employee_code': employee_code,
                    'errors': row_errors,
                    'raw_data': {k: str(v) if v is not None else '' for k, v in raw_data.items()},
                })
            else:
                # Attach raw_data for storage in BulkUploadRow
                cleaned['_raw'] = {k: str(v) if v is not None else '' for k, v in raw_data.items()}
                cleaned['_row_number'] = idx
                valid_rows.append(cleaned)

        wb.close()
        return valid_rows, error_rows
